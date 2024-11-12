# Q. Write a Python program to read data from an Excel file and use it in Selenium test cases.
#
# Example:
# Fetch a list of usernames and passwords from an Excel file and attempt to log into a website.
import time

# https://the-internet.herokuapp.com/login
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

file_path = r"./credentials.xlsx"
url = "https://the-internet.herokuapp.com/login"

# loads/opens the Excel workbook
workbook = openpyxl.load_workbook(file_path)

# move to active sheet
sheet = workbook.active


# Retrieve the login credentials (Data) from the Excel based on the URL
def retrieve_credentials(active_sheet):
    try:
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                if sheet.cell(i, j).value == url:
                    username = sheet.cell(i, 1).value
                    password = sheet.cell(i, 2).value
                    return username, password
    except Exception as e:
        print("Error finding the credentials for the requested URL. ", e)


# Open the URL in browser and login with the retrieved credentials
driver = webdriver.Chrome()
driver.get(url)
assert "Login Page" in driver.page_source

driver.find_element(By.NAME, "username").send_keys(retrieve_credentials(sheet)[0])
driver.find_element(By.NAME, "password").send_keys(retrieve_credentials(sheet)[1])

driver.find_element(By.TAG_NAME, "button").click()
assert "Welcome to the Secure Area." in driver.page_source

driver.find_element(By.XPATH, "//*[@class='button secondary radius']").click()

